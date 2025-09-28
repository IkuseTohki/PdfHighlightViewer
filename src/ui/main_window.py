"""アプリケーションのメインウィンドウとUIの振る舞いを定義します。"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import fitz
import os
from collections import defaultdict
from PIL import Image, ImageDraw
import io
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

from config.settings import AppSettings
from pdf import extractor, renderer

class MainWindow:
    """アプリケーションのメインウィンドウとUIロジックを管理するクラス。"""

    def __init__(self, root):
        """MainWindowを初期化します。

        Args:
            root (tk.Tk): Tkinterのルートウィンドウ。
        """
        self.root = root
        self.root.title("PDF Highlight Viewer")
        self.root.geometry("1200x800")

        # --- 状態変数 ---
        self.doc = None
        self.highlights = []
        self.page_images = [] # PhotoImageオブジェクトのガベージコレクションを防ぐための参照保持リスト
        self.current_page_num = -1
        self.current_highlight_rect = None
        self.scale = 1.0
        self.app_settings = AppSettings() # setting.iniから設定を読み込む
        self.export_format = tk.StringVar(value="png") # エクスポート形式

        self._setup_ui()
        self._apply_styles()

    def _setup_ui(self):
        """GUIウィジェットの初期化と配置を行います。"""
        # --- メニューバーの作成 ---
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # ファイルメニュー
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="ファイル", menu=file_menu)
        file_menu.add_command(label="PDFを開く...", command=self.open_pdf_file)
        file_menu.add_separator()
        file_menu.add_command(label="終了", command=self.root.quit)

        # エクスポートメニュー
        export_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="エクスポート", menu=export_menu)

        format_menu = tk.Menu(export_menu, tearoff=0)
        export_menu.add_cascade(label="エクスポート形式", menu=format_menu)
        format_menu.add_radiobutton(label="画像 (PNG)", variable=self.export_format, value="png")
        format_menu.add_radiobutton(label="PDF", variable=self.export_format, value="pdf")
        format_menu.add_radiobutton(label="Excel", variable=self.export_format, value="excel")

        export_menu.add_separator()
        export_menu.add_command(label="選択中のハイライトをエクスポート...", command=self.export_selected_highlight)
        export_menu.add_command(label="すべてのハイライトをエクスポート...", command=self.export_all_highlights)

        # --- メインフレーム ---
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=(0, 5))

        # open_pdf_fileはメニューに移動したが、ボタンも残しておく
        self.btn_open = ttk.Button(top_frame, text="PDFを開く", command=self.open_pdf_file)
        self.btn_open.pack(side=tk.LEFT)

        self.lbl_filepath = ttk.Label(top_frame, text="PDFファイルが選択されていません", anchor=tk.W)
        self.lbl_filepath.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

        # PanedWindowで左右のペインをリサイズ可能にする
        content_frame = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        content_frame.pack(fill=tk.BOTH, expand=True)

        # --- 左ペイン (検出リスト) ---
        list_frame = ttk.Frame(content_frame, padding=5)
        content_frame.add(list_frame, weight=1)

        self.list_label = ttk.Label(list_frame, text="検出された領域:")
        self.list_label.pack(anchor=tk.W)
        
        listbox_container = ttk.Frame(list_frame)
        listbox_container.pack(fill=tk.BOTH, expand=True, pady=(5,0))

        self.listbox = tk.Listbox(listbox_container)
        self.listbox_sby = ttk.Scrollbar(listbox_container, orient=tk.VERTICAL, command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=self.listbox_sby.set)
        self.listbox.bind("<<ListboxSelect>>", self.on_list_select)
        
        self.listbox_sby.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # --- 右ペイン (PDFビューア) ---
        viewer_frame = ttk.Frame(content_frame, padding=5)
        content_frame.add(viewer_frame, weight=4)

        preview_controls_frame = ttk.Frame(viewer_frame)
        preview_controls_frame.pack(fill=tk.X)

        self.viewer_label = ttk.Label(preview_controls_frame, text="PDFプレビュー:")
        self.viewer_label.pack(side=tk.LEFT, anchor=tk.W)

        # ズームコントロール
        zoom_frame = ttk.Frame(preview_controls_frame)
        zoom_frame.pack(side=tk.LEFT, padx=20)

        self.zoom_out_btn = ttk.Button(zoom_frame, text="-", command=self.zoom_out, width=2)
        self.zoom_out_btn.pack(side=tk.LEFT, padx=5)
        self.scale_label = ttk.Label(zoom_frame, text=f"{self.scale*100:.0f}%")
        self.scale_label.pack(side=tk.LEFT)
        self.zoom_in_btn = ttk.Button(zoom_frame, text="+", command=self.zoom_in, width=2)
        self.zoom_in_btn.pack(side=tk.LEFT, padx=5)

        # PDF表示用キャンバス
        canvas_container = ttk.Frame(viewer_frame)
        canvas_container.pack(fill=tk.BOTH, expand=True, pady=(5,0))

        self.canvas = tk.Canvas(canvas_container, bg="lightgray")
        self.canvas_vsb = ttk.Scrollbar(canvas_container, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas_hsb = ttk.Scrollbar(canvas_container, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.canvas_vsb.set, xscrollcommand=self.canvas_hsb.set)

        self.canvas_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas_hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def _apply_styles(self):
        """設定ファイルに基づいてUIのスタイル（フォントサイズなど）を適用します。"""
        font_size = self.app_settings.font_size
        default_font = ("TkDefaultFont", font_size)
        
        style = ttk.Style()
        style.configure(".", font=default_font) # すべてのttkウィジェットのデフォルトフォントを設定
        
        # ttkではない標準ウィジェットや、個別に設定が必要なウィジェット
        self.listbox.config(font=default_font)
        self.scale_label.config(font=default_font)

    def open_pdf_file(self):
        """ファイルダイアログを開き、選択されたPDFの処理を開始します。"""
        filepath = filedialog.askopenfilename(
            title="PDFファイルを選択",
            filetypes=[("PDF files", "*.pdf")]
        )
        if not filepath:
            return

        self._reset_state()
        self.lbl_filepath.config(text=filepath)

        try:
            self.doc = fitz.open(filepath)
            self.highlights = extractor.extract_colored_regions(self.doc, self.app_settings)
            self._populate_listbox()
            if self.doc.page_count > 0:
                self.show_page(0)
        except Exception as e:
            messagebox.showerror("エラー", f"PDFの処理中にエラーが発生しました:\n{e}")
            self.lbl_filepath.config(text=f"エラー: {e}")

    def _reset_state(self):
        """新しいPDFを開く前に、アプリケーションの状態をリセットします。"""
        self.listbox.delete(0, tk.END)
        self.canvas.delete("all")
        self.highlights = []
        self.page_images = []
        self.current_page_num = -1
        self.current_highlight_rect = None
        self.scale = 1.0
        if self.doc:
            self.doc.close()
            self.doc = None

    def _populate_listbox(self):
        """検出した領域のリストをリストボックスに表示します。"""
        self.listbox.delete(0, tk.END)
        for i, (page_num, rect) in enumerate(self.highlights):
            self.listbox.insert(tk.END, f"領域 {i + 1} (Page {page_num + 1})")

    def on_list_select(self, event):
        """リストボックスの項目が選択されたときに呼び出されるイベントハンドラです。"""
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            return
        
        selected_index = selection_indices[0]
        
        if 0 <= selected_index < len(self.highlights):
            page_num, rect = self.highlights[selected_index]
            self.current_highlight_rect = rect
            self.show_page(page_num, highlight_rect=rect)

    def show_page(self, page_num, highlight_rect=None):
        """指定されたページをキャンバスに表示し、指定領域をハイライトします。

        Args:
            page_num (int): 表示するページ番号。
            highlight_rect (fitz.Rect, optional): ハイライト表示する矩形領域。
                Defaults to None.
        """
        if not self.doc or not (0 <= page_num < self.doc.page_count):
            return

        self.current_page_num = page_num
        page = self.doc[page_num]
        # ページを指定倍率で画像にレンダリング
        img, photo = renderer.render_page_to_image(page, scale=self.scale)
        
        self.page_images.append(photo) # 参照を保持

        self.canvas.delete("all")
        self.canvas.create_image(0, 0, anchor=tk.NW, image=photo)
        self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

        # ハイライト用の赤枠を描画
        if highlight_rect:
            # 矩形の座標も表示倍率に合わせてスケールさせる
            self.canvas.create_rectangle(
                highlight_rect.x0 * self.scale,
                highlight_rect.y0 * self.scale,
                highlight_rect.x1 * self.scale,
                highlight_rect.y1 * self.scale,
                outline="red",
                width=3,
                tags="highlight" # タグを付けておくと後で削除しやすい
            )
            # 選択箇所が画面内に表示されるようにスクロール
            page_height = self.canvas.bbox(tk.ALL)[3]
            if page_height > 0:
                self.canvas.yview_moveto((highlight_rect.y0 * self.scale) / page_height)

    def export_selected_highlight(self):
        """選択されたハイライト領域を、指定された形式でエクスポートします。"""
        export_format = self.export_format.get()
        if export_format == "png":
            self._export_selected_highlight_as_image()
        elif export_format == "pdf":
            self._export_selected_highlight_as_pdf()
        elif export_format == "excel":
            self._export_selected_highlight_as_excel()
        else:
            messagebox.showinfo("未実装", f"{export_format}形式でのエクスポートはまだ実装されていません。")

    def export_all_highlights(self):
        """すべてのハイライト領域を、指定された形式でエクスポートします。"""
        export_format = self.export_format.get()
        if export_format == "png":
            self._export_all_highlights_as_image()
        elif export_format == "pdf":
            self._export_all_highlights_as_pdf()
        elif export_format == "excel":
            self._export_all_highlights_as_excel()
        else:
            messagebox.showinfo("未実装", f"{export_format}形式でのエクスポートはまだ実装されていません。")

    def _export_selected_highlight_as_image(self):
        """選択されたハイライト箇所を含むページ全体を画像としてエクスポートします。"""
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("エクスポート不可", "エクスポートする領域が選択されていません。")
            return

        selected_index = selection_indices[0]
        if not (0 <= selected_index < len(self.highlights)):
            return

        page_num, rect = self.highlights[selected_index]

        filepath = filedialog.asksaveasfilename(
            title="ページ画像を保存",
            defaultextension=".png",
            filetypes=[("PNG Image", "*.png"), ("JPEG Image", "*.jpg")]
        )
        if not filepath:
            return

        try:
            page = self.doc[page_num]
            dpi = 300
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)

            # ページ全体を高解像度でレンダリング
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            draw = ImageDraw.Draw(img)

            # 画像に合わせてハイライトの矩形座標をスケーリングして描画
            highlight_rect_on_image = rect * mat
            draw.rectangle(
                (highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1),
                outline="red",
                width=5 # 枠線を太くする
            )

            img.save(filepath)
            messagebox.showinfo("成功", f"ページ画像をエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"画像の保存中にエラーが発生しました:\n{e}")

    def _export_all_highlights_as_image(self):
        """すべてのハイライト箇所について、それぞれを含むページ全体を画像として保存します。"""
        if not self.highlights:
            messagebox.showwarning("エクスポート不可", "エクスポート対象の領域がありません。")
            return

        folder_path = filedialog.askdirectory(title="保存先のフォルダを選択")
        if not folder_path:
            return

        try:
            page_counters = defaultdict(int)
            exported_count = 0
            dpi = 300
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)

            for page_num, rect in self.highlights:
                page_counters[page_num] += 1
                counter = page_counters[page_num]
                
                filename = f"page-{page_num + 1}-{counter}.png"
                filepath = os.path.join(folder_path, filename)
                
                page = self.doc[page_num]
                
                # ページ全体をレンダリング
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                draw = ImageDraw.Draw(img)

                # ハイライト矩形を描画
                highlight_rect_on_image = rect * mat
                draw.rectangle(
                    (highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1),
                    outline="red",
                    width=5
                )

                img.save(filepath)
                exported_count += 1
            
            messagebox.showinfo("成功", f"{exported_count}個のページ画像をエクスポートしました。\nフォルダ: {folder_path}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"エクスポート中にエラーが発生しました:\n{e}")

    def _export_selected_highlight_as_pdf(self):
        """選択されたハイライト箇所を含むページを、1ページのPDFとしてエクスポートします。"""
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("エクスポート不可", "エクスポートする領域が選択されていません。")
            return

        selected_index = selection_indices[0]
        if not (0 <= selected_index < len(self.highlights)):
            return

        page_num, rect = self.highlights[selected_index]

        filepath = filedialog.asksaveasfilename(
            title="PDFとして保存",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")]
        )
        if not filepath:
            return

        try:
            new_doc = fitz.open() # 新しいPDFを作成
            new_doc.insert_pdf(self.doc, from_page=page_num, to_page=page_num)
            new_page = new_doc[0]
            new_page.draw_rect(rect, color=(1, 0, 0), width=1.5)
            new_doc.save(filepath)
            new_doc.close()
            messagebox.showinfo("成功", f"PDFをエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"PDFの保存中にエラーが発生しました:\n{e}")

    def _export_all_highlights_as_pdf(self):
        """すべてのハイライト箇所を、設定に応じて1つのPDFにまとめてエクスポートします。"""
        if not self.highlights:
            messagebox.showwarning("エクスポート不可", "エクスポート対象の領域がありません。")
            return

        filepath = filedialog.asksaveasfilename(
            title="PDFとして保存",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")]
        )
        if not filepath:
            return

        final_doc = fitz.open()
        try:
            if self.app_settings.pdf_export_mode == 'one_page': # A案
                for page_num, rect in self.highlights:
                    temp_doc = fitz.open()
                    temp_doc.insert_pdf(self.doc, from_page=page_num, to_page=page_num)
                    new_page = temp_doc[0]
                    new_page.draw_rect(rect, color=(1, 0, 0), width=1.5)
                    final_doc.insert_pdf(temp_doc)
                    temp_doc.close()
            
            elif self.app_settings.pdf_export_mode == 'merge': # B案
                highlights_by_page = defaultdict(list)
                for page_num, rect in self.highlights:
                    highlights_by_page[page_num].append(rect)
                
                sorted_pages = sorted(highlights_by_page.keys())

                for page_num in sorted_pages:
                    temp_doc = fitz.open()
                    temp_doc.insert_pdf(self.doc, from_page=page_num, to_page=page_num)
                    new_page = temp_doc[0]
                    for rect in highlights_by_page[page_num]:
                        new_page.draw_rect(rect, color=(1, 0, 0), width=1.5)
                    final_doc.insert_pdf(temp_doc)
                    temp_doc.close()

            final_doc.save(filepath)
            messagebox.showinfo("成功", f"{len(final_doc)}ページのPDFをエクスポートしました。\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"PDFのエクスポート中にエラーが発生しました:\n{e}")
        finally:
            final_doc.close()

    def _export_selected_highlight_as_excel(self):
        """選択されたハイライトを、新しい形式でExcelファイルとしてエクスポートします。"""
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("エクスポート不可", "エクスポートする領域が選択されていません。")
            return

        selected_index = selection_indices[0]
        if not (0 <= selected_index < len(self.highlights)):
            return

        filepath = filedialog.asksaveasfilename(
            title="Excelとして保存",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Highlight"
            ws.append(["No", "ページ画像", "ページ番号", "テキスト"])

            page_num, rect = self.highlights[selected_index]
            page = self.doc[page_num]

            # データを書き込む
            text = page.get_text("text", clip=rect).strip()
            ws.cell(row=2, column=1, value=1)
            ws.cell(row=2, column=3, value=page_num + 1)
            ws.cell(row=2, column=4, value=text)

            # 画像の描画
            scale = self.app_settings.excel_image_scale
            mat = fitz.Matrix(scale, scale)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            draw = ImageDraw.Draw(img)
            highlight_rect_on_image = rect * mat
            draw.rectangle(
                (highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1),
                outline="red", width=3
            )
            
            img_path = io.BytesIO()
            img.save(img_path, format="PNG")
            img_path.seek(0)
            
            img_for_excel = OpenpyxlImage(img_path)
            ws.add_image(img_for_excel, "B2")

            # スタイル調整
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = img_for_excel.width * 0.14 # 近似値
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 50
            ws.row_dimensions[2].height = img_for_excel.height * 0.75 # 近似値

            wb.save(filepath)
            messagebox.showinfo("成功", f"Excelファイルをエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"Excelファイルのエクスポート中にエラーが発生しました:\n{e}")

    def _export_all_highlights_as_excel(self):
        """すべてのハイライトを、セル結合を使用した形式で1つのExcelファイルにエクスポートします。"""
        if not self.highlights:
            messagebox.showwarning("エクスポート不可", "エクスポート対象の領域がありません。")
            return

        filepath = filedialog.asksaveasfilename(
            title="Excelとして保存",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Highlights"
            ws.append(["No", "ページ画像", "ページ番号", "テキスト"])

            highlights_by_page = defaultdict(list)
            for page_num, rect in self.highlights:
                highlights_by_page[page_num].append(rect)
            
            sorted_pages = sorted(highlights_by_page.keys())

            current_row = 2
            highlight_no = 1
            max_image_width = 0

            for page_num in sorted_pages:
                rects = highlights_by_page[page_num]
                num_highlights = len(rects)
                start_row = current_row
                end_row = start_row + num_highlights - 1

                # セルを結合 (B列: 画像, C列: ページ番号)
                if num_highlights > 1:
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                    ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                
                # ページ番号を書き込み
                ws.cell(row=start_row, column=3, value=page_num + 1)
                # 中央揃え
                ws.cell(row=start_row, column=3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

                # ページ画像を作成
                page = self.doc[page_num]
                scale = self.app_settings.excel_image_scale
                mat = fitz.Matrix(scale, scale)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                draw = ImageDraw.Draw(img)
                for r in rects:
                    highlight_rect_on_image = r * mat
                    draw.rectangle(
                        (highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1),
                        outline="red", width=3
                    )
                
                img_path = io.BytesIO()
                img.save(img_path, format="PNG")
                img_path.seek(0)
                img_for_excel = OpenpyxlImage(img_path)
                ws.add_image(img_for_excel, f"B{start_row}")
                if img_for_excel.width > max_image_width:
                    max_image_width = img_for_excel.width

                # 行の高さを調整
                total_height_points = img_for_excel.height * 0.75
                height_per_row = total_height_points / num_highlights
                for i in range(num_highlights):
                    ws.row_dimensions[start_row + i].height = height_per_row

                # 各ハイライトの情報を書き込み
                for rect in rects:
                    text = page.get_text("text", clip=rect).strip()
                    ws.cell(row=current_row, column=1, value=highlight_no)
                    ws.cell(row=current_row, column=4, value=text)
                    highlight_no += 1
                    current_row += 1

            # 全体のスタイル調整
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = max_image_width * 0.14
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 50
            
            wb.save(filepath)
            messagebox.showinfo("成功", f"{highlight_no - 1}個のハイライトをExcelファイルにエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"Excelファイルのエクスポート中にエラーが発生しました:\n{e}")

    def zoom_in(self):
        """表示倍率を上げて再描画します。"""
        self.scale += 0.1
        self.scale_label.config(text=f"{self.scale*100:.0f}%")
        self.show_page(self.current_page_num, highlight_rect=self.current_highlight_rect)

    def zoom_out(self):
        """表示倍率を下げて再描画します。"""
        if self.scale > 0.2:
            self.scale -= 0.1
            self.scale_label.config(text=f"{self.scale*100:.0f}%")
            self.show_page(self.current_page_num, highlight_rect=self.current_highlight_rect)