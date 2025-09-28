"""PDFのハイライト情報を様々な形式でエクスポートする機能を提供します。"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import fitz
import os
from collections import defaultdict
from PIL import Image, ImageDraw
import io
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

class Exporter:
    """エクスポート処理を実行するクラス。"""

    def __init__(self, doc, highlights, listbox, app_settings):
        """Exporterを初期化します。

        Args:
            doc (fitz.Document): 操作対象のPDFドキュメント。
            highlights (list): 抽出されたハイライト情報のリスト。
            listbox (tk.Listbox): UIのリストボックスウィジェット。
            app_settings (AppSettings): アプリケーションの設定オブジェクト。
        """
        self.doc = doc
        self.highlights = highlights
        self.listbox = listbox
        self.app_settings = app_settings

    def export_selected(self, export_format):
        """選択されたハイライト領域を、指定された形式でエクスポートします。"""
        if not self.doc or not self.highlights:
            messagebox.showwarning("エクスポート不可", "エクスポート対象のPDFが開かれていません。")
            return

        if export_format == "png":
            self._export_selected_highlight_as_image()
        elif export_format == "pdf":
            self._export_selected_highlight_as_pdf()
        elif export_format == "excel":
            self._export_selected_highlight_as_excel()
        else:
            messagebox.showinfo("未実装", f"{export_format}形式でのエクスポートはまだ実装されていません。")

    def export_all(self, export_format):
        """すべてのハイライト領域を、指定された形式でエクスポートします。"""
        if not self.doc or not self.highlights:
            messagebox.showwarning("エクスポート不可", "エクスポート対象のPDFが開かれていません。")
            return

        if export_format == "png":
            self._export_all_highlights_as_image()
        elif export_format == "pdf":
            self._export_all_highlights_as_pdf()
        elif export_format == "excel":
            self._export_all_highlights_as_excel()
        else:
            messagebox.showinfo("未実装", f"{export_format}形式でのエクスポートはまだ実装されていません。")

    # --- Private Image Export Methods ---
    def _export_selected_highlight_as_image(self):
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("エクスポート不可", "エクスポートする領域が選択されていません。")
            return
        selected_index = selection_indices[0]
        if not (0 <= selected_index < len(self.highlights)):
            return
        page_num, rect = self.highlights[selected_index]
        filepath = filedialog.asksaveasfilename(title="ページ画像を保存", defaultextension=".png", filetypes=[("PNG Image", "*.png"), ("JPEG Image", "*.jpg")])
        if not filepath:
            return
        try:
            page = self.doc[page_num]
            dpi = 300
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            draw = ImageDraw.Draw(img)
            highlight_rect_on_image = rect * mat
            draw.rectangle((highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1), outline="red", width=5)
            img.save(filepath)
            messagebox.showinfo("成功", f"ページ画像をエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"画像の保存中にエラーが発生しました:\n{e}")

    def _export_all_highlights_as_image(self):
        folder_path = filedialog.askdirectory(title="保存先のフォルダを選択")
        if not folder_path:
            return
        try:
            page_counters = defaultdict(int)
            exported_count = 0
            dpi = 300
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            for page_num, rect in self.highlights:
                page_counters[page_num] += 1
                counter = page_counters[page_num]
                filename = f"page-{page_num + 1}-{counter}.png"
                filepath = os.path.join(folder_path, filename)
                page = self.doc[page_num]
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                draw = ImageDraw.Draw(img)
                highlight_rect_on_image = rect * mat
                draw.rectangle((highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1), outline="red", width=5)
                img.save(filepath)
                exported_count += 1
            messagebox.showinfo("成功", f"{exported_count}個のページ画像をエクスポートしました。\nフォルダ: {folder_path}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"エクスポート中にエラーが発生しました:\n{e}")

    # --- Private PDF Export Methods ---
    def _export_selected_highlight_as_pdf(self):
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("エクスポート不可", "エクスポートする領域が選択されていません。")
            return
        selected_index = selection_indices[0]
        if not (0 <= selected_index < len(self.highlights)):
            return
        page_num, rect = self.highlights[selected_index]
        filepath = filedialog.asksaveasfilename(title="PDFとして保存", defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not filepath:
            return
        try:
            new_doc = fitz.open()
            new_doc.insert_pdf(self.doc, from_page=page_num, to_page=page_num)
            new_page = new_doc[0]
            new_page.draw_rect(rect, color=(1, 0, 0), width=1.5)
            new_doc.save(filepath)
            new_doc.close()
            messagebox.showinfo("成功", f"PDFをエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"PDFの保存中にエラーが発生しました:\n{e}")

    def _export_all_highlights_as_pdf(self):
        filepath = filedialog.asksaveasfilename(title="PDFとして保存", defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not filepath:
            return
        final_doc = fitz.open()
        try:
            if self.app_settings.pdf_export_mode == 'one_page':
                for page_num, rect in self.highlights:
                    temp_doc = fitz.open()
                    temp_doc.insert_pdf(self.doc, from_page=page_num, to_page=page_num)
                    new_page = temp_doc[0]
                    new_page.draw_rect(rect, color=(1, 0, 0), width=1.5)
                    final_doc.insert_pdf(temp_doc)
                    temp_doc.close()
            elif self.app_settings.pdf_export_mode == 'merge':
                highlights_by_page = defaultdict(list)
                for page_num, rect in self.highlights:
                    highlights_by_page[page_num].append(rect)
                sorted_pages = sorted(highlights_by_page.keys())
                for page_num in sorted_pages:
                    temp_doc = fitz.open()
                    temp_doc.insert_pdf(self.doc, from_page=page_num, to_page=page_num)
                    new_page = temp_doc[0]
                    for rect in highlights_by_page[page_num]:
                        new_page.draw_rect(rect, color=(1, 0, 0), width=1.5)
                    final_doc.insert_pdf(temp_doc)
                    temp_doc.close()
            final_doc.save(filepath)
            messagebox.showinfo("成功", f"{len(final_doc)}ページのPDFをエクスポートしました。\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"PDFのエクスポート中にエラーが発生しました:\n{e}")
        finally:
            final_doc.close()

    # --- Private Excel Export Methods ---
    def _export_selected_highlight_as_excel(self):
        selection_indices = self.listbox.curselection()
        if not selection_indices:
            messagebox.showwarning("エクスポート不可", "エクスポートする領域が選択されていません。")
            return
        selected_index = selection_indices[0]
        if not (0 <= selected_index < len(self.highlights)):
            return
        filepath = filedialog.asksaveasfilename(title="Excelとして保存", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not filepath:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Highlight"
            ws.append(["No", "ページ画像", "ページ番号", "テキスト"])
            page_num, rect = self.highlights[selected_index]
            page = self.doc[page_num]
            text = page.get_text("text", clip=rect).strip()
            ws.cell(row=2, column=1, value=1)
            ws.cell(row=2, column=3, value=page_num + 1)
            ws.cell(row=2, column=4, value=text)
            scale = self.app_settings.excel_image_scale
            mat = fitz.Matrix(scale, scale)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            draw = ImageDraw.Draw(img)
            highlight_rect_on_image = rect * mat
            draw.rectangle((highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1), outline="red", width=3)
            img_path = io.BytesIO()
            img.save(img_path, format="PNG")
            img_path.seek(0)
            img_for_excel = OpenpyxlImage(img_path)
            ws.add_image(img_for_excel, "B2")
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = img_for_excel.width * 0.14
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 50
            ws.row_dimensions[2].height = img_for_excel.height * 0.75
            wb.save(filepath)
            messagebox.showinfo("成功", f"Excelファイルをエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"Excelファイルのエクスポート中にエラーが発生しました:\n{e}")

    def _export_all_highlights_as_excel(self):
        filepath = filedialog.asksaveasfilename(title="Excelとして保存", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not filepath:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Highlights"
            ws.append(["No", "ページ画像", "ページ番号", "テキスト"])
            highlights_by_page = defaultdict(list)
            for page_num, rect in self.highlights:
                highlights_by_page[page_num].append(rect)
            sorted_pages = sorted(highlights_by_page.keys())
            current_row = 2
            highlight_no = 1
            max_image_width = 0
            for page_num in sorted_pages:
                rects = highlights_by_page[page_num]
                num_highlights = len(rects)
                start_row = current_row
                end_row = start_row + num_highlights - 1
                if num_highlights > 1:
                    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                    ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
                ws.cell(row=start_row, column=3, value=page_num + 1)
                ws.cell(row=start_row, column=3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                page = self.doc[page_num]
                scale = self.app_settings.excel_image_scale
                mat = fitz.Matrix(scale, scale)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                draw = ImageDraw.Draw(img)
                for r in rects:
                    highlight_rect_on_image = r * mat
                    draw.rectangle((highlight_rect_on_image.x0, highlight_rect_on_image.y0, highlight_rect_on_image.x1, highlight_rect_on_image.y1), outline="red", width=3)
                img_path = io.BytesIO()
                img.save(img_path, format="PNG")
                img_path.seek(0)
                img_for_excel = OpenpyxlImage(img_path)
                ws.add_image(img_for_excel, f"B{start_row}")
                if img_for_excel.width > max_image_width:
                    max_image_width = img_for_excel.width
                total_height_points = img_for_excel.height * 0.75
                height_per_row = total_height_points / num_highlights
                for i in range(num_highlights):
                    ws.row_dimensions[start_row + i].height = height_per_row
                for rect in rects:
                    text = page.get_text("text", clip=rect).strip()
                    ws.cell(row=current_row, column=1, value=highlight_no)
                    ws.cell(row=current_row, column=4, value=text)
                    highlight_no += 1
                    current_row += 1
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = max_image_width * 0.14
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 50
            wb.save(filepath)
            messagebox.showinfo("成功", f"{highlight_no - 1}個のハイライトをExcelファイルにエクスポートしました:\n{filepath}")
        except Exception as e:
            messagebox.showerror("エクスポートエラー", f"Excelファイルのエクスポート中にエラーが発生しました:\n{e}")
